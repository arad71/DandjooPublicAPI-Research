import abc
import csv
import datetime
from typing import Generator

import chardet
from openpyxl import load_workbook
from shapefile import Reader


class BaseFileReaderInterface(metaclass=abc.ABCMeta):
    """
    This class is a Python psuedo interface with the intention that all implementations can be used by the calling
    class consistently without relying specifics of the file type being read.
    """

    def __init__(self, file_path: str):
        self._file_path = file_path

    @classmethod
    def __subclasshook__(cls, subclass):
        """
        Check subclass has implemented all methods of BaseReader.

        :param subclass: subclass being checked
        :return: true if all methods are implemented
        """
        base_method_list = [func for func in dir(BaseFileReaderInterface) if
                            callable(getattr(BaseFileReaderInterface, func)) and not func.startswith("__")]

        return all([hasattr(subclass, method_name) and callable(getattr(subclass, method_name)) for method_name in
                    base_method_list])

    @abc.abstractmethod
    def get_data_sample(self) -> dict:
        """
        This method will read the first item of the data to provide a sample of the data.

        :return: A dictionary with keys being the data field names and values being field values of the data.
        """
        raise NotImplementedError

    @property
    @abc.abstractmethod
    def records(self) -> Generator[dict, None, None]:
        """
        This method will return an iterable of each record in the file.

        :return: A generator object that will yield records (dictionaries with key / values of the data).
        """
        raise NotImplementedError


class CSVFileReader(BaseFileReaderInterface):
    _encoding = ''

    def __init__(self, file_path: str):
        super().__init__(file_path)

        with open(self._file_path, 'rb') as csvfile:
            csv_contents = csvfile.read()

            encoding_result = chardet.detect(csv_contents)

            self._encoding = encoding_result['encoding']

    def get_data_sample(self) -> dict:
        with open(self._file_path, encoding=self._encoding) as csvfile:
            csvreader = csv.reader(csvfile)

            header_row = next(csvreader)

            try:
                first_data_row = next(csvreader)
            except StopIteration:
                # assign empty values when there are no data rows
                first_data_row = ['' for _ in header_row]

            return dict(zip(header_row, first_data_row))

    @property
    def records(self) -> Generator[dict, None, None]:
        with open(self._file_path, encoding=self._encoding) as csvfile:
            csvreader = csv.reader(csvfile)

            header_row = next(csvreader)

            for data_row in csvreader:
                yield dict(zip(header_row, data_row))


class ExcelFileReader(BaseFileReaderInterface):
    def get_data_sample(self) -> dict:
        workbook = load_workbook(self._file_path, True)

        # content is expected to be in first worksheet
        worksheet = workbook.worksheets[0]

        # note: worksheet[1] and worksheet[2] are the first rows
        header_row = [cell.value for cell in worksheet[1]]

        # first row data is a list of values or empty list if the first row has no data
        # note: max_row being None indicated openpyxl cannot detect the number of rows but we assume it's more than one
        if worksheet.max_row is None or worksheet.max_row > 1:
            first_data_row = [cell.value for cell in worksheet[2]]
        else:
            first_data_row = []

        return dict(zip(header_row, first_data_row))

    @property
    def records(self) -> Generator[dict, None, None]:
        workbook = load_workbook(self._file_path, True)

        # content is expected to be in first worksheet
        worksheet = workbook.worksheets[0]

        # note: worksheet[1] is the first row
        header_row = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(2):
            data_row = [cell.value for cell in row]

            yield dict(zip(header_row, data_row))


class ShapeFileReader(BaseFileReaderInterface):
    EXCLUDED_ATTRIBUTES = ['DeletionFlag']

    def get_data_sample(self) -> dict:
        shapefile = Reader(self._file_path)

        attributes_names = [field[0] for field in shapefile.fields if field[0] not in self.EXCLUDED_ATTRIBUTES]

        # first feature data is a list of values or empty list if the first row has no data
        if shapefile.numRecords > 0:
            # get attribute values of first feature (which doesn't include DeletionFlag) and convert any dates to isoformat
            first_record_data = [value.isoformat() if isinstance(value, datetime.date) else value for value in
                                 shapefile.record(0)]
        else:
            first_record_data = []

        return dict(zip(attributes_names, first_record_data))

    @property
    def records(self) -> Generator[dict, None, None]:
        shapefile = Reader(self._file_path)

        attributes_names = [field[0] for field in shapefile.fields if field[0] != 'DeletionFlag']

        for record_attributes in shapefile.records(attributes_names):
            record_data = [value.isoformat() if isinstance(value, datetime.date) else value for value in
                           record_attributes]

            yield dict(zip(attributes_names, record_data))
