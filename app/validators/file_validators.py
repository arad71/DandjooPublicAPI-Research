import csv
import os

from zipfile import ZipFile, BadZipFile

import chardet
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from shapefile import Reader, ShapefileException

from app.validators.utils import is_int_string, is_float_string

FILE_SIZE_ERROR = 'File Size Error'
FILE_FORMAT_ERROR = 'File Format Error'
FILE_HEADER_ERROR = 'File Header Error'
GENERAL_ERROR = 'General Error'


class FileValidator:
    DEFAULT_MAX_FILE_SIZE = 524288000

    def __init__(self, file_path: str):
        self._file_path = file_path

        self._errors = {}

    def validate(self):
        self.validate_file_size()

    @property
    def is_valid(self):
        return not bool(self._errors)

    @property
    def errors(self):
        return self._errors

    def has_error(self, error_type):
        return error_type in self._errors

    def has_errors(self, error_types):
        return any(self.has_error(error_type) for error_type in error_types)

    def _add_error(self, error_type, error_detail):
        if self.has_error(error_type):
            self._errors[error_type].append(error_detail)
        else:
            self._errors[error_type] = [error_detail]

    def validate_file_size(self, max_size_in_bytes: int = DEFAULT_MAX_FILE_SIZE):
        if os.stat(self._file_path).st_size > max_size_in_bytes:
            self._add_error(FILE_SIZE_ERROR,
                            f'File size is larger than maximum allowable size of {max_size_in_bytes}')


class CSVFileValidator(FileValidator):
    ACCEPTED_ENCODINGS = ['utf-8', 'ascii', 'ISO-8859-1']
    _encoding = ''

    def __init__(self, file_path: str):
        super().__init__(file_path)

    def validate(self):
        super().validate()

        self.validate_file_character_encoding()

        if not self.has_errors([FILE_FORMAT_ERROR, GENERAL_ERROR]):
            self.validate_header_row_not_empty()

        if not self.has_errors([FILE_FORMAT_ERROR, FILE_HEADER_ERROR, GENERAL_ERROR]):
            self.validate_header_row_content()

    def validate_file_character_encoding(self):
        with open(self._file_path, 'rb') as csvfile:
            csv_contents = csvfile.read()

            try:
                encoding_result = chardet.detect(csv_contents)
            except TypeError as te:
                self._add_error(FILE_FORMAT_ERROR, str(te))
            else:
                self._encoding = encoding_result['encoding']
                if self._encoding not in self.ACCEPTED_ENCODINGS:
                    self._add_error(FILE_FORMAT_ERROR,
                                    f"File character encoding must be one of {', '.join(self.ACCEPTED_ENCODINGS)}; "
                                    f"{self._encoding} detected")

    def validate_header_row_not_empty(self):
        with open(self._file_path, encoding=self._encoding) as csvfile:
            csvreader = csv.reader(csvfile)

            try:
                header_row = next(csvreader)

                # note: worksheet[1] is the first row
                if all(cell is None for cell in header_row):
                    self._add_error(FILE_HEADER_ERROR, 'The header row must contain data')
            except Exception as e:
                self._add_error(GENERAL_ERROR, str(e))

    def validate_header_row_content(self):
        # open file again to reset reader
        with open(self._file_path, encoding=self._encoding) as csvfile:
            csvreader = csv.reader(csvfile)

            try:
                header_row = next(csvreader)

                progressive_header_values = []
                for cell in header_row:
                    if cell in progressive_header_values:
                        self._add_error(FILE_HEADER_ERROR, 'All header cells must contain unique values')
                        break

                    progressive_header_values.append(cell)

                for cell in header_row:
                    if is_int_string(cell) or is_float_string(cell):
                        self._add_error(FILE_HEADER_ERROR, 'All header cells must be a string (text)')
                        break
            except Exception as e:
                self._add_error(GENERAL_ERROR, str(e))


class ExcelFileValidator(FileValidator):
    def __init__(self, file_path: str):
        super().__init__(file_path)

    def validate(self):
        super().validate()

        self.validate_file_format()

        if not self.has_error(FILE_FORMAT_ERROR):
            self.validate_header_row_not_empty()

        if not self.has_errors([FILE_FORMAT_ERROR, FILE_HEADER_ERROR]):
            self.validate_header_row_content()

    def validate_file_format(self):
        try:
            load_workbook(self._file_path, True)
        except InvalidFileException as ife:
            raise self._add_error(FILE_FORMAT_ERROR, str(ife))
        except BadZipFile as bzf:
            self._add_error(FILE_FORMAT_ERROR, str(bzf))

    def validate_header_row_not_empty(self):
        workbook = load_workbook(self._file_path, True)

        # content is expected to be in first worksheet
        worksheet = workbook.worksheets[0]

        # note: worksheet[1] is the first row
        if all(cell.value is None for cell in worksheet[1]):
            self._add_error(FILE_HEADER_ERROR, 'The header row must contain data')

    def validate_header_row_content(self):
        workbook = load_workbook(self._file_path, True)

        # content is expected to be in first worksheet
        worksheet = workbook.worksheets[0]

        progressive_header_values = []
        for cell in worksheet[1]:
            if cell.value in progressive_header_values:
                self._add_error(FILE_HEADER_ERROR, 'All header cells must contain unique values')
                break

            progressive_header_values.append(cell.value)

        for cell in worksheet[1]:
            if type(cell.value) is not str:
                self._add_error(FILE_HEADER_ERROR, 'All header cells must be a string (text)')
                break


class ShapeFileValidator(FileValidator):
    REQUIRED_SHAPEFILE_FILE_TYPES = ['.shp', '.shx', '.dbf']
    OPTIONAL_SHAPEFILE_FILE_TYPES = ['.prj', '.xml', '.sbn', '.sbx', '.cpg']

    def __init__(self, file_path: str, expected_geometry_types: list):
        super().__init__(file_path)

        self.expected_geometry_types = expected_geometry_types

    def validate(self):
        super().validate()

        self.validate_zip_file()

        if not self.has_error(FILE_FORMAT_ERROR):
            self.validate_shapefile()

        if self.expected_geometry_types is not None and not self.has_error(FILE_FORMAT_ERROR):
            self.validate_shapefile_type(self.expected_geometry_types)

    def validate_zip_file(self):
        try:
            zip_file = ZipFile(self._file_path)
        except BadZipFile as bzf:
            self._add_error(FILE_FORMAT_ERROR, str(bzf))
        else:
            if zip_file.testzip() is not None:
                self._add_error(FILE_FORMAT_ERROR, 'Zipfile is invalid')

            for file_type in self.REQUIRED_SHAPEFILE_FILE_TYPES:
                if len([f for f in zip_file.namelist() if f.endswith(file_type)]) != 1:
                    self._add_error(FILE_FORMAT_ERROR, f"Zipfile must contain exactly one '{file_type}' file")

            allowed_files_types = self.REQUIRED_SHAPEFILE_FILE_TYPES + self.OPTIONAL_SHAPEFILE_FILE_TYPES

            contains_invalid_file_type = False
            for filename in zip_file.namelist():
                if os.path.splitext(filename)[1].lower() not in allowed_files_types:
                    contains_invalid_file_type = True

            if contains_invalid_file_type:
                self._add_error(FILE_FORMAT_ERROR, 'Zipfile must only contain files with shapefile-related '
                                                   f"extensions ({'/'.join(allowed_files_types)})")

    def validate_shapefile(self):
        try:
            reader = Reader(self._file_path)
        except ShapefileException as sfe:
            self._add_error(FILE_FORMAT_ERROR, str(sfe))
        else:
            # create list of attributes that don't include the default fields, DeletionFlag and OBJECTID
            mappable_attributes = [field for field in reader.fields if field[0] not in ['DeletionFlag', 'OBJECTID']]
            if len(mappable_attributes) < 2:
                self._add_error(FILE_FORMAT_ERROR,
                                'Shapefile is missing DBF file or has less than 2 non-default attributes')

    def validate_shapefile_type(self, shape_types):
        shapefile = Reader(self._file_path)

        if shapefile.shapeTypeName not in shape_types:
            self._add_error(FILE_FORMAT_ERROR, f"Shapefile must contain {'/'.join(shape_types)} based geometry")
